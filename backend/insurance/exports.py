import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm

HEADERS = ["Claim #", "Patient", "Hospital #", "Insurer", "Member #", "Status", "Claimed (KES)", "Approved (KES)", "Submitted", "Created"]


def _row(c):
    return [
        c.claim_number,
        c.patient.full_name,
        c.patient.hospital_number,
        c.policy.insurer.name,
        c.policy.member_number,
        c.get_status_display(),
        f"{c.total_claimed:,.2f}",
        f"{c.total_approved:,.2f}",
        c.submitted_at.strftime("%Y-%m-%d %H:%M") if c.submitted_at else "",
        c.created_at.strftime("%Y-%m-%d %H:%M"),
    ]


def export_claims_xlsx(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Insurance Claims"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")

    for c in queryset:
        ws.append(_row(c))

    for i, width in enumerate([16, 24, 14, 20, 16, 18, 14, 14, 18, 18], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="insurance_claims.xlsx"'
    wb.save(response)
    return response


def export_claims_pdf(queryset):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1 * cm, rightMargin=1 * cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph("Insurance Claims", styles["Title"]), Spacer(1, 12)]

    data = [HEADERS] + [_row(c) for c in queryset]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="insurance_claims.pdf"'
    return response
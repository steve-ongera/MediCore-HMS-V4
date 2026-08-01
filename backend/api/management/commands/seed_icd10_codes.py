"""
Seeds the icd10_codes table from the bundled XLSX data file.

Existing codes are left untouched — only codes not already present in the
database are inserted, so this is safe to re-run at any time (e.g. after
pulling in an updated spreadsheet with more codes).

Requires openpyxl (pip install openpyxl) to read the .xlsx file.

Usage:
    python manage.py seed_icd10_codes
    python manage.py seed_icd10_codes --dry-run
    python manage.py seed_icd10_codes --file /path/to/other_codes.xlsx
"""
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from api.models import ICD10Code

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

DEFAULT_XLSX_PATH = Path(__file__).resolve().parent / "data" / "icd10_codes.xlsx"

REQUIRED_COLUMNS = {"code", "description", "category"}


class Command(BaseCommand):
    help = "Seed ICD-10 codes from CSV, skipping codes that already exist."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            type=str,
            default=None,
            help=f"Path to the XLSX file to load (default: {DEFAULT_XLSX_PATH}).",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Show what would be created without writing to the database.",
        )
        parser.add_argument(
            "--batch-size",
            type=int,
            default=500,
            help="Number of rows per bulk_create batch (default: 500).",
        )

    def handle(self, *args, **options):
        if load_workbook is None:
            raise CommandError(
                "openpyxl is required to read .xlsx files. Install it with: "
                "pip install openpyxl"
            )

        xlsx_path = Path(options["file"]) if options["file"] else DEFAULT_XLSX_PATH
        dry_run = options["dry_run"]
        batch_size = options["batch_size"]

        if not xlsx_path.exists():
            raise CommandError(f"XLSX file not found: {xlsx_path}")

        rows = self._read_xlsx(xlsx_path)
        self.stdout.write(f"Read {len(rows)} rows from {xlsx_path}")

        existing_codes = set(ICD10Code.objects.values_list("code", flat=True))
        self.stdout.write(f"{len(existing_codes)} ICD-10 codes already in the database")

        seen_in_file = set()
        new_objects = []
        skipped_existing = 0
        skipped_duplicate_in_file = 0
        skipped_invalid = 0

        for row in rows:
            code = str(row.get("code") or "").strip()
            description = str(row.get("description") or "").strip()
            category = str(row.get("category") or "").strip()

            if not code or not description:
                skipped_invalid += 1
                continue

            if len(code) > 10:
                self.stdout.write(
                    self.style.WARNING(f"Skipping {code!r}: exceeds 10-char code limit")
                )
                skipped_invalid += 1
                continue

            if code in existing_codes:
                skipped_existing += 1
                continue

            if code in seen_in_file:
                skipped_duplicate_in_file += 1
                continue

            seen_in_file.add(code)
            new_objects.append(
                ICD10Code(code=code, description=description, category=category)
            )

        self.stdout.write(f"New codes to insert: {len(new_objects)}")
        self.stdout.write(f"Skipped (already in DB): {skipped_existing}")
        if skipped_duplicate_in_file:
            self.stdout.write(f"Skipped (duplicate within file): {skipped_duplicate_in_file}")
        if skipped_invalid:
            self.stdout.write(f"Skipped (invalid row): {skipped_invalid}")

        if dry_run:
            self.stdout.write(self.style.WARNING("Dry run — no changes written."))
            return

        if not new_objects:
            self.stdout.write(self.style.SUCCESS("Nothing to insert. Database is up to date."))
            return

        with transaction.atomic():
            created = ICD10Code.objects.bulk_create(
                new_objects,
                batch_size=batch_size,
                ignore_conflicts=True,  # extra safety net against race conditions
            )

        self.stdout.write(
            self.style.SUCCESS(f"Successfully seeded {len(created)} ICD-10 codes.")
        )

    def _read_xlsx(self, path: Path) -> list[dict]:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise CommandError(f"XLSX file is empty: {path}")

        headers = [str(h).strip().lower() if h is not None else "" for h in header_row]
        if not REQUIRED_COLUMNS.issubset(set(headers)):
            raise CommandError(
                f"XLSX must have columns {REQUIRED_COLUMNS}, found {headers}"
            )

        col_index = {h: i for i, h in enumerate(headers)}
        rows = []
        for values in rows_iter:
            if values is None or all(v is None for v in values):
                continue  # skip blank rows
            row = {
                "code": values[col_index["code"]],
                "description": values[col_index["description"]],
                "category": values[col_index["category"]],
            }
            rows.append(row)

        wb.close()
        return rows